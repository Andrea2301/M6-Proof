import openpyxl
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Employees"

# Headers
headers = [
    "Document", "FirstName", "LastName", "Email", 
    "Phone", "Address", "BirthDate", "Salary", 
    "HireDate", "ProfessionalProfile", "Position", "Department"
]

for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# Sample data
sample_data = [
    ["12345678", "John", "Doe", "john.doe@example.com", 
     "555-1234", "123 Main St", "1990-01-15", "50000", 
     "2023-01-01", "Experienced developer", "Desarrollador", "TI"],
    ["87654321", "Jane", "Smith", "jane.smith@example.com", 
     "555-5678", "456 Oak Ave", "1985-05-20", "60000", 
     "2022-06-15", "Senior analyst", "Analista", "RRHH"]
]

for row_idx, data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

wb.save("wwwroot/sample/employees-template.xlsx")
print("Excel template created")
