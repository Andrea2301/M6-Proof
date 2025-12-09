import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Employees"
headers = ["Document", "FirstName", "LastName", "Email", "Phone", "Address", "BirthDate", "Salary", "HireDate", "ProfessionalProfile"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)
data = [
    ["12345678", "John", "Doe", "john.doe@example.com", "555-1234", "123 Main St", "1990-01-15", 50000, "2023-01-01", "Experienced developer"],
    ["87654321", "Jane", "Smith", "jane.smith@example.com", "555-5678", "456 Oak Ave", "1985-05-20", 60000, "2022-06-15", "Senior analyst"]
]
for row_idx, row_data in enumerate(data, 2):
    for col_idx, cell_data in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=cell_data)
wb.save("wwwroot/sample/employees-template.xlsx")
print("Excel template created")
